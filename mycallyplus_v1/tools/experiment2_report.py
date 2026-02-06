#!/usr/bin/env python3
"""
Generate consolidated tables and figures for experiment2 results (VM & RK3588).
Outputs are written to: mycallyplus_v1/experiment/experiment2实验总表/

Data sources (read-only):
- mycallyplus_v1/experiment/experiment2虚拟机实验结果汇总/*/run_results_*.txt (and summary.json if present)
- mycallyplus_v1/experiment/experiment2（rk3588）实验结果/*/run_results_*.txt (and summary.json if present)
"""
import json
import math
import os
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import math

# Excel formatting helpers (optional at runtime)
try:
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    get_column_letter = None

BASE = Path(__file__).resolve().parent.parent
SRC_DIRS = {
    "vm": BASE / "experiment" / "experiment2虚拟机实验结果汇总",
    "rk3588": BASE / "experiment" / "experiment2（rk3588）实验结果",
}
OUT_DIR = BASE / "experiment" / "experiment2实验总表"

# ---------------- Parsing helpers ----------------

def parse_dir_meta(dir_name: str) -> Dict:
    # Examples: 2026-01-31T06-46-17-08-00_ws100_r50_c4, 2026-02-02T08-29-28+00-00_ws50_r48
    m = re.search(r"ws(\d+)_r(\d+)(?:_c(\d+))?", dir_name)
    ws = int(m.group(1)) if m else None
    r = int(m.group(2)) if m else None
    c = int(m.group(3)) if m and m.group(3) else None
    return {"work_scale": ws, "repeats": r, "cores_per_task": c}


def parse_summary_json(path: Path) -> Dict:
    data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    ws = data.get("work_scale")
    r = data.get("repeats")
    c = data.get("cores_per_task")

    def stats_from(times: List[float]) -> Dict:
        s = pd.Series(times)
        return {
            "n": len(times),
            "mean": s.mean(),
            "median": s.median(),
            "std": s.std(ddof=1),
            "min": s.min(),
            "max": s.max(),
            "p25": s.quantile(0.25),
            "p75": s.quantile(0.75),
        }

    baseline_times = data["baseline"]["times_s"]
    prio_times = data["prio"]["times_s"]
    b_stats = stats_from(baseline_times)
    p_stats = stats_from(prio_times)

    runs = data.get("runs", [])
    long_rows = []
    pair_rows = []
    # Build long & paired
    by_iter = {}
    for r_entry in runs:
        iter_idx = int(r_entry.get("iter", 0))
        variant = r_entry.get("program")
        t = float(r_entry.get("time_s", r_entry.get("wall_s", 0.0)))
        rc = r_entry.get("returncode", 0)
        long_rows.append({"run_idx": iter_idx, "variant": variant, "time_s": t, "rc": rc})
        by_iter.setdefault(iter_idx, {})[variant] = t
    for idx, pair in by_iter.items():
        if "baseline" in pair and "prio" in pair:
            base_t = pair["baseline"]
            prio_t = pair["prio"]
            pair_rows.append({
                "run_idx": idx,
                "baseline_s": base_t,
                "prio_s": prio_t,
                "delta_s": base_t - prio_t,
                "ratio": base_t / prio_t if prio_t else math.nan,
            })

    return {
        "ws": ws,
        "repeats": r,
        "cores": c,
        "baseline_times": baseline_times,
        "prio_times": prio_times,
        "b_stats": b_stats,
        "p_stats": p_stats,
        "long_rows": long_rows,
        "pair_rows": pair_rows,
    }


def parse_run_results_txt(path: Path) -> Dict:
    txt = path.read_text(encoding="utf-8", errors="ignore")
    meta = parse_dir_meta(path.parent.name)
    ws, r, c = meta.get("work_scale"), meta.get("repeats"), meta.get("cores_per_task")
    # extract times list from "times_s=[...]"
    def extract_times(section: str) -> List[float]:
        m = re.search(r"times_s=\[(.*?)\]", section, re.S)
        if not m:
            return []
        body = m.group(1)
        nums = re.findall(r"[0-9]+\.?[0-9]*", body)
        return [float(x) for x in nums]

    # split baseline / prio sections
    b_sec = re.search(r"=== BASELINE STATS.*?(?=== PRIO STATS|\Z)", txt, re.S)
    p_sec = re.search(r"=== PRIO STATS.*?(?=== LOGS|\Z)", txt, re.S)
    b_times = extract_times(b_sec.group(0)) if b_sec else []
    p_times = extract_times(p_sec.group(0)) if p_sec else []
    def stats_from(times: List[float]) -> Dict:
        if not times:
            return {k: math.nan for k in ["n","mean","median","std","min","max","p25","p75"]}
        s = pd.Series(times)
        return {
            "n": len(times),
            "mean": s.mean(),
            "median": s.median(),
            "std": s.std(ddof=1),
            "min": s.min(),
            "max": s.max(),
            "p25": s.quantile(0.25),
            "p75": s.quantile(0.75),
        }
    b_stats = stats_from(b_times)
    p_stats = stats_from(p_times)

    # try to extract per-run table lines: pattern: idx baseline prio delta
    long_rows = []
    pair_rows = []
    # In many files, first lines like: 35\t0.811328\t0.625942\t-0.185386
    for line in txt.splitlines():
        if re.match(r"^\d+\t", line.strip()):
            parts = line.strip().split("\t")
            if len(parts) >= 3:
                try:
                    idx = int(parts[0])
                    b = float(parts[1])
                    p = float(parts[2])
                except Exception:
                    continue
                long_rows.append({"run_idx": idx-1, "variant": "baseline", "time_s": b, "rc": 0})
                long_rows.append({"run_idx": idx-1, "variant": "prio", "time_s": p, "rc": 0})
                pair_rows.append({
                    "run_idx": idx-1,
                    "baseline_s": b,
                    "prio_s": p,
                    "delta_s": b - p,
                    "ratio": b / p if p else math.nan,
                })
    return {
        "ws": ws,
        "repeats": r,
        "cores": c,
        "baseline_times": b_times,
        "prio_times": p_times,
        "b_stats": b_stats,
        "p_stats": p_stats,
        "long_rows": long_rows,
        "pair_rows": pair_rows,
    }


def collect_all():
    exp_rows = []
    long_rows = []
    pair_rows = []
    for platform, base in SRC_DIRS.items():
        if not base.exists():
            continue
        for exp_dir in sorted(base.iterdir()):
            if not exp_dir.is_dir():
                continue
            meta = parse_dir_meta(exp_dir.name)
            ws, r, c = meta.get("work_scale"), meta.get("repeats"), meta.get("cores_per_task")
            summary_path = exp_dir / "summary.json"
            if summary_path.exists():
                parsed = parse_summary_json(summary_path)
            else:
                run_files = list(exp_dir.glob("run_results_*.txt"))
                if not run_files:
                    continue
                parsed = parse_run_results_txt(run_files[0])
            # build exp row
            b = parsed["b_stats"]
            p = parsed["p_stats"]
            cv_b = b["std"] / b["mean"] if b["mean"] else math.nan
            cv_p = p["std"] / p["mean"] if p["mean"] else math.nan
            exp_rows.append({
                "platform": platform,
                "exp_id": exp_dir.name,
                "ws": parsed["ws"] or ws,
                "repeats": parsed["repeats"] or r,
                "cores": parsed["cores"] or c,
                "baseline_mean": b["mean"],
                "baseline_median": b["median"],
                "baseline_std": b["std"],
                "baseline_min": b["min"],
                "baseline_max": b["max"],
                "baseline_p25": b["p25"],
                "baseline_p75": b["p75"],
                "baseline_cv": cv_b,
                "prio_mean": p["mean"],
                "prio_median": p["median"],
                "prio_std": p["std"],
                "prio_min": p["min"],
                "prio_max": p["max"],
                "prio_p25": p["p25"],
                "prio_p75": p["p75"],
                "prio_cv": cv_p,
                "n_baseline": b["n"],
                "n_prio": p["n"],
                "speedup": b["mean"] / p["mean"] if p["mean"] else math.nan,
                "improve_pct": (b["mean"] - p["mean"]) / b["mean"] * 100 if b["mean"] else math.nan,
            })
            # append long rows
            for row in parsed["long_rows"]:
                row = dict(row)
                row.update({"platform": platform, "exp_id": exp_dir.name, "ws": parsed["ws"] or ws})
                long_rows.append(row)
            for row in parsed["pair_rows"]:
                row = dict(row)
                row.update({"platform": platform, "exp_id": exp_dir.name, "ws": parsed["ws"] or ws})
                pair_rows.append(row)
    return pd.DataFrame(exp_rows), pd.DataFrame(long_rows), pd.DataFrame(pair_rows)


# ---------------- Plot helpers ----------------

def save_fig(fig, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, bbox_inches="tight", dpi=200)
    plt.close(fig)


def plot_per_exp(exp_id: str, df_long: pd.DataFrame, df_pair: pd.DataFrame, platform: str):
    out_dir = OUT_DIR / "figures" / "per_exp" / platform / exp_id
    out_dir.mkdir(parents=True, exist_ok=True)
    # save data slices
    df_long.to_csv(out_dir / "times_long.csv", index=False)
    df_pair.to_csv(out_dir / "paired.csv", index=False)

    # 1) distribution baseline vs prio
    fig, ax = plt.subplots(figsize=(4,4))
    sns.boxplot(data=df_long, x="variant", y="time_s", ax=ax)
    sns.stripplot(data=df_long, x="variant", y="time_s", ax=ax, color="0.3", alpha=0.5, jitter=0.2)
    ax.set_title(f"{exp_id}\n时间分布")
    ax.set_ylabel("time (s)")
    save_fig(fig, out_dir / "dist_box.png")

    # 2) delta histogram
    if not df_pair.empty:
        fig, ax = plt.subplots(figsize=(4,3))
        sns.histplot(df_pair["delta_s"], bins=20, kde=True, ax=ax)
        ax.axvline(0, color="red", ls="--")
        ax.set_title(f"{exp_id}\n baseline - prio")
        ax.set_xlabel("delta (s)")
        save_fig(fig, out_dir / "delta_hist.png")

    # 3) run series
    fig, ax = plt.subplots(figsize=(5,3))
    sns.lineplot(data=df_long, x="run_idx", y="time_s", hue="variant", marker="o", ax=ax)
    ax.set_title(f"{exp_id}\n按运行序列")
    ax.set_ylabel("time (s)")
    save_fig(fig, out_dir / "run_series.png")


def plot_global(df_exp: pd.DataFrame):
    gdir = OUT_DIR / "figures" / "global"
    gdir.mkdir(parents=True, exist_ok=True)
    df_exp.to_csv(OUT_DIR / "tables" / "experiment_index.csv", index=False)

    # speedup vs ws
    fig, ax = plt.subplots(figsize=(6,4))
    sns.lineplot(data=df_exp, x="ws", y="speedup", hue="platform", marker="o", ax=ax)
    ax.set_xscale('log')
    ax.set_title("Speedup vs work_scale")
    ax.set_ylabel("speedup (baseline/prio)")
    save_fig(fig, gdir / "speedup_vs_ws.png")

    # mean time vs ws (baseline/prio)
    melted = df_exp.melt(id_vars=["platform","ws"], value_vars=["baseline_mean","prio_mean"], var_name="variant", value_name="mean_s")
    fig, ax = plt.subplots(figsize=(6,4))
    sns.lineplot(data=melted, x="ws", y="mean_s", hue="platform", style="variant", marker="o", ax=ax)
    ax.set_xscale('log')
    ax.set_ylabel("mean time (s)")
    ax.set_title("Mean runtime vs work_scale")
    save_fig(fig, gdir / "mean_vs_ws.png")

    # CV vs ws
    cv_melt = df_exp.melt(id_vars=["platform","ws"], value_vars=["baseline_cv","prio_cv"], var_name="variant", value_name="cv")
    fig, ax = plt.subplots(figsize=(6,4))
    sns.lineplot(data=cv_melt, x="ws", y="cv", hue="platform", style="variant", marker="o", ax=ax)
    ax.set_xscale('log')
    ax.set_ylabel("CV")
    ax.set_title("Coefficient of variation vs work_scale")
    save_fig(fig, gdir / "cv_vs_ws.png")


def _autoformat_ws(ws_sheet):
    if get_column_letter is None:
        return
    # Freeze header and add autofilter
    ws_sheet.freeze_panes = "A2"
    ws_sheet.auto_filter.ref = ws_sheet.dimensions
    # Auto width
    for col_cells in ws_sheet.columns:
        max_len = 0
        col = col_cells[0].column
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws_sheet.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)


def save_pretty_tables(df_exp, df_long, df_pair, df_plat):
    tables_dir = OUT_DIR / "tables"
    tables_dir.mkdir(parents=True, exist_ok=True)

    # Rounded copies for readability
    def round_df(df):
        return df.copy().applymap(lambda x: round(x, 3) if isinstance(x, (int, float, float)) else x)

    df_exp_r = df_exp.copy()
    num_cols = df_exp_r.select_dtypes(include=["float", "int"]).columns
    df_exp_r[num_cols] = df_exp_r[num_cols].round(3)

    df_plat_r = df_plat.copy()
    num_cols = df_plat_r.select_dtypes(include=["float", "int"]).columns
    df_plat_r[num_cols] = df_plat_r[num_cols].round(3)

    df_pair_r = df_pair.copy()
    num_cols = df_pair_r.select_dtypes(include=["float", "int"]).columns
    df_pair_r[num_cols] = df_pair_r[num_cols].round(4)

    df_long_r = df_long.copy()
    num_cols = df_long_r.select_dtypes(include=["float", "int"]).columns
    df_long_r[num_cols] = df_long_r[num_cols].round(4)

    # Write Excel with multiple sheets
    xlsx_path = tables_dir / "summary.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_exp_r.to_excel(writer, sheet_name="experiment_index", index=False)
        df_plat_r.to_excel(writer, sheet_name="platform_summary", index=False)
        df_long_r.to_excel(writer, sheet_name="run_level_long", index=False)
        df_pair_r.to_excel(writer, sheet_name="pair_level", index=False)
        # combined_summary is same as platform_summary here
        df_plat_r.to_excel(writer, sheet_name="combined_summary", index=False)

        # Apply formatting
        for ws in writer.book.worksheets:
            _autoformat_ws(ws)

    # Optional: Parquet (skip if engine missing)
    try:
        df_exp.to_parquet(tables_dir / "experiment_index.parquet", index=False)
        df_long.to_parquet(tables_dir / "run_level_long.parquet", index=False)
        df_pair.to_parquet(tables_dir / "pair_level.parquet", index=False)
        df_plat.to_parquet(tables_dir / "platform_summary.parquet", index=False)
    except Exception:
        pass

    # Human-friendly markdown overview (top rows + schema)
    md_path = tables_dir / "overview.md"
    lines = []
    lines.append("# experiment2 汇总表（美化版概览）\n")
    lines.append("## experiment_index（前 10 行）\n")
    lines.append(df_exp_r.head(10).to_markdown(index=False))
    lines.append("\n## platform_summary\n")
    lines.append(df_plat_r.to_markdown(index=False))
    lines.append("\n数据文件：\n")
    lines.append(f"- Excel: `{xlsx_path.relative_to(BASE)}`\n- Parquet: 同目录下 *.parquet（若环境支持会生成）\n- 原始 CSV 仍保留在 `tables/`")
    md_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------- Main ----------------

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "tables").mkdir(exist_ok=True)

    df_exp, df_long, df_pair = collect_all()
    # Save tables
    df_long.to_csv(OUT_DIR / "tables" / "run_level_long.csv", index=False)
    df_pair.to_csv(OUT_DIR / "tables" / "pair_level.csv", index=False)

    # platform summary
    plat_rows = []
    for (plat, ws), sub in df_exp.groupby(["platform","ws"]):
        b_mean = sub["baseline_mean"].mean()
        p_mean = sub["prio_mean"].mean()
        plat_rows.append({
            "platform": plat,
            "ws": ws,
            "baseline_mean": b_mean,
            "prio_mean": p_mean,
            "speedup": b_mean / p_mean if p_mean else math.nan,
            "improve_pct": (b_mean - p_mean) / b_mean * 100 if b_mean else math.nan,
            "n_experiments": len(sub),
        })
    df_plat = pd.DataFrame(plat_rows)
    df_plat.to_csv(OUT_DIR / "tables" / "platform_summary.csv", index=False)

    # combined summary (just clone df_plat but marked)
    df_plat.to_csv(OUT_DIR / "tables" / "combined_summary.csv", index=False)

    # Pretty outputs (Excel, Parquet, Markdown overview)
    save_pretty_tables(df_exp, df_long, df_pair, df_plat)

    # per-exp plots
    for (plat, exp_id), sub_long in df_long.groupby(["platform","exp_id"]):
        sub_pair = df_pair[(df_pair.platform==plat) & (df_pair.exp_id==exp_id)]
        plot_per_exp(exp_id, sub_long, sub_pair, plat)

    # global plots
    plot_global(df_exp)

if __name__ == "__main__":
    main()
